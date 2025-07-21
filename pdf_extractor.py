#!/usr/bin/env python3
"""
PDF Data Extractor Tool
A secure, local application for extracting specific data from PDF files and exporting to Excel.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import os
import re
import threading
from datetime import datetime
import logging
from typing import List, Dict, Tuple, Optional
import pytesseract
from PIL import Image
import io
import platform
import sys
from insurance_extractor_mode import EnhancedInsuranceExtractor
from idp_enhanced_extractor import IDPInsuranceExtractor

class PDFDataExtractor:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF Data Extractor - IDP/ICR Enhanced")
        self.root.geometry("900x750")
        
        # Configure logging for debugging
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        self.logger = logging.getLogger(__name__)
        
        # Initialize variables
        self.selected_files = []
        self.search_terms = []
        self.extracted_data = []
        self.is_processing = False
        self.insurance_mode = False
        self.idp_mode = False
        
        # Initialize both extractors
        self.insurance_extractor = EnhancedInsuranceExtractor()
        self.idp_extractor = IDPInsuranceExtractor()
        
        # Setup OCR for bundled environment
        self.setup_ocr_environment()
        
        self.setup_ui()
    
    def setup_ocr_environment(self):
        """Configure OCR paths for both development and bundled environments"""
        try:
            # Check if running as a PyInstaller bundle
            if hasattr(sys, '_MEIPASS'):
                # Running as PyInstaller bundle
                bundle_dir = sys._MEIPASS
                self.logger.info(f"Running as bundle, base directory: {bundle_dir}")
                
                # Set tesseract executable path (platform-specific)
                if platform.system() == "Windows":
                    tesseract_path = os.path.join(bundle_dir, 'tesseract', 'tesseract.exe')
                else:
                    tesseract_path = os.path.join(bundle_dir, 'tesseract', 'tesseract')
                
                if os.path.exists(tesseract_path):
                    pytesseract.pytesseract.tesseract_cmd = tesseract_path
                    self.logger.info(f"Set tesseract path: {tesseract_path}")
                
                # Set tessdata directory
                tessdata_path = os.path.join(bundle_dir, 'tessdata')
                if os.path.exists(tessdata_path):
                    os.environ['TESSDATA_PREFIX'] = tessdata_path
                    self.logger.info(f"Set TESSDATA_PREFIX: {tessdata_path}")
            else:
                # Running in development mode - use system tesseract
                self.logger.info("Running in development mode, using system tesseract")
                
                # Try to find system tesseract if not already configured
                try:
                    import subprocess
                    if platform.system() == "Windows":
                        # Windows uses 'where' command
                        result = subprocess.run(["where", "tesseract"], capture_output=True, text=True, shell=True)
                    else:
                        # Unix-like systems use 'which' command
                        result = subprocess.run(["which", "tesseract"], capture_output=True, text=True)
                    
                    if result.returncode == 0:
                        tesseract_path = result.stdout.strip().split('\n')[0]  # Take first result
                        pytesseract.pytesseract.tesseract_cmd = tesseract_path
                        self.logger.info(f"Found system tesseract: {tesseract_path}")
                except Exception as e:
                    self.logger.warning(f"Could not locate system tesseract: {e}")
            
            # Test OCR functionality
            try:
                version = pytesseract.get_tesseract_version()
                self.logger.info(f"OCR ready - Tesseract version: {version}")
            except Exception as e:
                self.logger.error(f"OCR setup failed: {e}")
                
        except Exception as e:
            self.logger.error(f"Error setting up OCR environment: {e}")
        
    def setup_ui(self):
        """Create the user interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="PDF Data Extractor", 
                               font=('Helvetica', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # File selection section
        ttk.Label(main_frame, text="1. Select PDF Files:", font=('Helvetica', 12, 'bold')).grid(
            row=1, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))
        
        file_frame = ttk.Frame(main_frame)
        file_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)
        
        ttk.Button(file_frame, text="Browse Files", 
                  command=self.browse_files).grid(row=0, column=1, padx=(5, 0))
        
        self.files_listbox = tk.Listbox(file_frame, height=6, selectmode=tk.EXTENDED)
        self.files_listbox.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # Scrollbar for listbox
        files_scrollbar = ttk.Scrollbar(file_frame, orient=tk.VERTICAL, 
                                       command=self.files_listbox.yview)
        files_scrollbar.grid(row=0, column=2, sticky=(tk.N, tk.S))
        self.files_listbox.configure(yscrollcommand=files_scrollbar.set)
        
        # Clear files button
        ttk.Button(file_frame, text="Clear", 
                  command=self.clear_files).grid(row=1, column=1, padx=(5, 0), pady=(5, 0))
        
        # Search terms section
        ttk.Label(main_frame, text="2. Enter Search Terms:", font=('Helvetica', 12, 'bold')).grid(
            row=3, column=0, columnspan=3, sticky=tk.W, pady=(20, 5))
        
        ttk.Label(main_frame, text="Enter search terms (one per line):").grid(
            row=4, column=0, columnspan=3, sticky=tk.W)
        
        self.search_text = scrolledtext.ScrolledText(main_frame, height=8, width=70)
        self.search_text.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 10))
        
        # Search options
        options_frame = ttk.LabelFrame(main_frame, text="Search Options", padding="5")
        options_frame.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.case_sensitive_var = tk.BooleanVar()
        self.whole_words_var = tk.BooleanVar()
        self.regex_var = tk.BooleanVar()
        self.context_var = tk.BooleanVar(value=True)
        self.force_ocr_var = tk.BooleanVar()
        self.auto_ocr_var = tk.BooleanVar(value=True)
        self.insurance_mode_var = tk.BooleanVar()
        self.idp_mode_var = tk.BooleanVar()
        
        # Insurance Mode - prominently placed at the top
        insurance_frame = ttk.Frame(options_frame)
        insurance_frame.grid(row=0, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 5))
        
        self.insurance_check = ttk.Checkbutton(insurance_frame, text="🏢 Insurance Mode (Auto-extract all 15 fields)", 
                                             variable=self.insurance_mode_var, command=self.toggle_insurance_mode)
        self.insurance_check.grid(row=0, column=0, sticky=tk.W)
        
        ttk.Label(insurance_frame, text="→ Extracts: Policy No, Names, Premiums, GST, Vehicle details, etc.", 
                 font=('Helvetica', 9), foreground='darkgreen').grid(row=1, column=0, sticky=tk.W)
        
        # IDP Mode - Enhanced accuracy mode
        idp_frame = ttk.Frame(options_frame)
        idp_frame.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(5, 10))
        
        self.idp_check = ttk.Checkbutton(idp_frame, text="🎯 IDP Mode (100% Accuracy + Show All Data)", 
                                        variable=self.idp_mode_var, command=self.toggle_idp_mode)
        self.idp_check.grid(row=0, column=0, sticky=tk.W)
        
        ttk.Label(idp_frame, text="→ AI-powered extraction + Shows unmatched data for manual review", 
                 font=('Helvetica', 9), foreground='darkblue').grid(row=1, column=0, sticky=tk.W)
        
        # Separator
        ttk.Separator(options_frame, orient='horizontal').grid(row=2, column=0, columnspan=4, 
                                                              sticky=(tk.W, tk.E), pady=(5, 10))
        
        # Regular search options
        self.case_check = ttk.Checkbutton(options_frame, text="Case Sensitive", 
                                         variable=self.case_sensitive_var)
        self.case_check.grid(row=3, column=0, sticky=tk.W, padx=(0, 15))
        
        self.words_check = ttk.Checkbutton(options_frame, text="Whole Words Only", 
                                          variable=self.whole_words_var)
        self.words_check.grid(row=3, column=1, sticky=tk.W, padx=(0, 15))
        
        self.regex_check = ttk.Checkbutton(options_frame, text="Regular Expressions", 
                                          variable=self.regex_var)
        self.regex_check.grid(row=3, column=2, sticky=tk.W, padx=(0, 15))
        
        self.context_check = ttk.Checkbutton(options_frame, text="Include Context", 
                                           variable=self.context_var)
        self.context_check.grid(row=3, column=3, sticky=tk.W)
        
        # OCR Options (fourth row)
        self.auto_ocr_check = ttk.Checkbutton(options_frame, text="Auto OCR (scanned PDFs)", 
                                            variable=self.auto_ocr_var)
        self.auto_ocr_check.grid(row=4, column=0, sticky=tk.W, padx=(0, 15), pady=(5, 0))
        
        self.force_ocr_check = ttk.Checkbutton(options_frame, text="Force OCR (all PDFs)", 
                                             variable=self.force_ocr_var)
        self.force_ocr_check.grid(row=4, column=1, sticky=tk.W, padx=(0, 15), pady=(5, 0))
        
        # Context length
        ttk.Label(options_frame, text="Context chars:").grid(row=4, column=2, sticky=tk.W, pady=(5, 0))
        self.context_length = tk.StringVar(value="100")
        self.context_entry = ttk.Entry(options_frame, textvariable=self.context_length, width=10)
        self.context_entry.grid(row=4, column=3, sticky=tk.W, pady=(5, 0))
        
        # OCR Help text
        ocr_help = ttk.Label(options_frame, text="💡 OCR extracts text from image-based/scanned PDFs", 
                            font=('Helvetica', 9))
        ocr_help.grid(row=5, column=0, columnspan=2, sticky=tk.W, pady=(3, 0))
        
        # Process button
        self.process_button = ttk.Button(main_frame, text="Extract Data", 
                                        command=self.start_processing, style='Accent.TButton')
        self.process_button.grid(row=7, column=0, columnspan=3, pady=20)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, 
                                           maximum=100)
        self.progress_bar.grid(row=8, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Status label
        self.status_label = ttk.Label(main_frame, text="Ready to process files...")
        self.status_label.grid(row=9, column=0, columnspan=3)
        
        # Results frame
        results_frame = ttk.LabelFrame(main_frame, text="Results", padding="5")
        results_frame.grid(row=10, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        results_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(10, weight=1)
        
        self.results_text = scrolledtext.ScrolledText(results_frame, height=10)
        self.results_text.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 5))
        
        # Export button
        self.export_button = ttk.Button(results_frame, text="Export to Excel", 
                                       command=self.export_to_excel, state=tk.DISABLED)
        self.export_button.grid(row=1, column=0, pady=(5, 0))
        
        # Clear results button
        ttk.Button(results_frame, text="Clear Results", 
                  command=self.clear_results).grid(row=1, column=1, padx=(10, 0), pady=(5, 0))
        
    def browse_files(self):
        """Open file dialog to select PDF files"""
        filetypes = [("PDF files", "*.pdf"), ("All files", "*.*")]
        files = filedialog.askopenfilenames(title="Select PDF Files", filetypes=filetypes)
        
        if files:
            self.selected_files.extend(files)
            # Update listbox
            self.files_listbox.delete(0, tk.END)
            for file in self.selected_files:
                self.files_listbox.insert(tk.END, os.path.basename(file))
            
            self.status_label.config(text=f"{len(self.selected_files)} files selected")
    
    def clear_files(self):
        """Clear selected files"""
        self.selected_files.clear()
        self.files_listbox.delete(0, tk.END)
        self.status_label.config(text="Ready to process files...")
    
    def clear_results(self):
        """Clear results display"""
        self.results_text.delete(1.0, tk.END)
        self.extracted_data.clear()
        self.export_button.config(state=tk.DISABLED)
    
    def toggle_insurance_mode(self):
        """Toggle between insurance mode and normal search mode"""
        insurance_mode = self.insurance_mode_var.get()
        
        if insurance_mode:
            # Disable IDP mode if it was enabled
            self.idp_mode_var.set(False)
            
            # Insurance Mode ON - disable search options and pre-fill search terms
            self.search_text.config(state=tk.DISABLED)
            self.case_check.config(state=tk.DISABLED)
            self.words_check.config(state=tk.DISABLED)
            self.regex_check.config(state=tk.DISABLED)
            self.context_check.config(state=tk.DISABLED)
            self.context_entry.config(state=tk.DISABLED)
            
            # Pre-fill with insurance terms
            self.search_text.config(state=tk.NORMAL)
            self.search_text.delete(1.0, tk.END)
            insurance_terms = """Policy no.
Insured name
Insurer name
Engine no.
Chassis no.
Cheque no.
Cheque date
Bank name
Net own damage premium amount
Net liability premium amount
Total premium amount
GST amount
Gross premium paid
Car model
Body type"""
            self.search_text.insert(1.0, insurance_terms)
            self.search_text.config(state=tk.DISABLED)
            
            self.status_label.config(text="Insurance Mode: Auto-extract all 15 fields enabled")
        else:
            # Insurance Mode OFF - enable search options
            self.search_text.config(state=tk.NORMAL)
            self.case_check.config(state=tk.NORMAL)
            self.words_check.config(state=tk.NORMAL)
            self.regex_check.config(state=tk.NORMAL)
            self.context_check.config(state=tk.NORMAL)
            self.context_entry.config(state=tk.NORMAL)
            
            # Clear search terms
            self.search_text.delete(1.0, tk.END)
            
            self.status_label.config(text="Normal Mode: Enter search terms manually")
    
    def toggle_idp_mode(self):
        """Toggle between IDP mode and normal search mode"""
        idp_mode = self.idp_mode_var.get()
        
        if idp_mode:
            # Disable Insurance mode if it was enabled
            self.insurance_mode_var.set(False)
            
            # IDP Mode ON - disable search options and pre-fill search terms
            self.search_text.config(state=tk.DISABLED)
            self.case_check.config(state=tk.DISABLED)
            self.words_check.config(state=tk.DISABLED)
            self.regex_check.config(state=tk.DISABLED)
            self.context_check.config(state=tk.DISABLED)
            self.context_entry.config(state=tk.DISABLED)
            
            # Pre-fill with IDP terms
            self.search_text.config(state=tk.NORMAL)
            self.search_text.delete(1.0, tk.END)
            idp_terms = """Policy no.
Insured name
Insurer name
Engine no.
Chassis no.
Cheque no.
Cheque date
Bank name
Net own damage premium amount
Net liability premium amount
Total premium amount
GST amount
Gross premium paid
Car model
Body type"""
            self.search_text.insert(1.0, idp_terms)
            self.search_text.config(state=tk.DISABLED)
            
            self.status_label.config(text="IDP Mode: 100% Accuracy + Unmatched Data enabled")
        else:
            # IDP Mode OFF - enable search options
            self.search_text.config(state=tk.NORMAL)
            self.case_check.config(state=tk.NORMAL)
            self.words_check.config(state=tk.NORMAL)
            self.regex_check.config(state=tk.NORMAL)
            self.context_check.config(state=tk.NORMAL)
            self.context_entry.config(state=tk.NORMAL)
            
            # Clear search terms
            self.search_text.delete(1.0, tk.END)
            
            self.status_label.config(text="Normal Mode: Enter search terms manually")
    
    def start_processing(self):
        """Start the PDF processing in a separate thread"""
        if not self.selected_files:
            messagebox.showerror("Error", "Please select PDF files first.")
            return
        
        # Check if insurance mode, IDP mode, or normal mode
        if not self.insurance_mode_var.get() and not self.idp_mode_var.get():
            # Normal mode - need search terms
            search_terms_text = self.search_text.get(1.0, tk.END).strip()
            if not search_terms_text:
                messagebox.showerror("Error", "Please enter search terms or enable Insurance/IDP Mode.")
                return
        
        # Parse search terms (only for normal mode)
        if not self.insurance_mode_var.get() and not self.idp_mode_var.get():
            self.search_terms = [term.strip() for term in search_terms_text.split('\n') 
                                if term.strip()]
            
            if not self.search_terms:
                messagebox.showerror("Error", "Please enter at least one search term.")
                return
        else:
            # Insurance or IDP mode - we'll use the respective extractors
            self.search_terms = []  # Not used in these modes
        
        # Disable process button and start processing
        self.process_button.config(state=tk.DISABLED)
        self.export_button.config(state=tk.DISABLED)
        self.is_processing = True
        
        # Start processing thread
        thread = threading.Thread(target=self.process_files)
        thread.daemon = True
        thread.start()
    
    def process_files(self):
        """Process all selected PDF files"""
        try:
            self.extracted_data.clear()
            total_files = len(self.selected_files)
            
            for i, file_path in enumerate(self.selected_files):
                if not self.is_processing:
                    break
                
                self.root.after(0, lambda: self.status_label.config(
                    text=f"Processing: {os.path.basename(file_path)} ({i+1}/{total_files})"))
                
                # Update progress bar
                progress = (i / total_files) * 100
                self.root.after(0, lambda p=progress: self.progress_var.set(p))
                
                try:
                    file_results = self.extract_from_pdf(file_path)
                    self.extracted_data.extend(file_results)
                except Exception as e:
                    self.logger.error(f"Error processing {file_path}: {str(e)}")
                    self.root.after(0, lambda err=str(e), file=file_path: 
                                   self.results_text.insert(tk.END, 
                                   f"Error processing {os.path.basename(file)}: {err}\n"))
            
            # Complete
            self.root.after(0, lambda: self.progress_var.set(100))
            self.root.after(0, self.processing_complete)
            
        except Exception as e:
            self.logger.error(f"Processing error: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("Error", f"Processing failed: {str(e)}"))
            self.root.after(0, self.reset_ui)
    
    def extract_from_pdf(self, file_path: str) -> List[Dict]:
        """Extract data from a single PDF file with OCR support"""
        results = []
        filename = os.path.basename(file_path)
        
        try:
            with pdfplumber.open(file_path) as pdf:
                full_text = ""
                page_texts = {}
                total_chars = 0
                
                # First, try normal text extraction (unless forcing OCR)
                if not self.force_ocr_var.get():
                    self.root.after(0, lambda: self.status_label.config(
                        text=f"Extracting text from {filename}..."))
                    
                    for page_num, page in enumerate(pdf.pages, 1):
                        page_text = page.extract_text()
                        if page_text:
                            page_texts[page_num] = page_text
                            full_text += f"\n--- Page {page_num} ---\n" + page_text
                            total_chars += len(page_text.strip())
                
                # Check if we need OCR (low text content or forced OCR)
                avg_chars_per_page = total_chars / max(len(pdf.pages), 1) if pdf.pages else 0
                need_ocr = (
                    self.force_ocr_var.get() or 
                    (self.auto_ocr_var.get() and avg_chars_per_page < 50)
                )
                
                if need_ocr:
                    self.root.after(0, lambda: self.status_label.config(
                        text=f"OCR processing {filename} - may take longer..."))
                    
                    # Perform OCR extraction
                    ocr_texts = self.extract_text_with_ocr(pdf, filename)
                    
                    # Use OCR results if we got more text, or if forcing OCR
                    if self.force_ocr_var.get() or len(''.join(ocr_texts.values())) > total_chars:
                        page_texts = ocr_texts
                        full_text = ""
                        for page_num in sorted(page_texts.keys()):
                            full_text += f"\n--- Page {page_num} ---\n" + page_texts[page_num]
                        
                        self.logger.info(f"Using OCR results for {filename}")
                        extraction_method = "OCR"
                    else:
                        self.logger.info(f"Normal text extraction sufficient for {filename}")
                        extraction_method = "Normal"
                else:
                    extraction_method = "Normal"
                
                # Handle insurance mode vs IDP mode vs normal search mode
                if self.idp_mode_var.get():
                    # IDP Mode - use comprehensive extraction with 100% coverage
                    idp_results = self.idp_extractor.extract_with_100_percent_coverage(full_text, filename)
                    
                    # Convert IDP results to results format
                    for field_key, field_data in idp_results['required_fields'].items():
                        if field_data['best_match']:
                            result = {
                                'filename': filename,
                                'search_term': field_data['field_name'],
                                'page': 'Multiple',  # IDP fields can be on any page
                                'context': field_data['best_match'].context,
                                'match': field_data['best_match'].value,
                                'extraction_method': f"IDP-{field_data['best_match'].method}",
                                'confidence': field_data['best_match'].confidence,
                                'validation_score': field_data['best_match'].validation_score,
                                'idp_field': field_key,
                                'field_type': field_data.get('field_type', 'unknown')
                            }
                            results.append(result)
                    
                    # Add unmatched data for manual review
                    unmatched = idp_results['unmatched_candidates']
                    for amount in unmatched.get('unmatched_monetary_amounts', [])[:5]:  # Limit to 5
                        result = {
                            'filename': filename,
                            'search_term': 'UNMATCHED Monetary Amount',
                            'page': 'Multiple',
                            'context': amount.get('context', ''),
                            'match': amount['value'],
                            'extraction_method': 'IDP-Unmatched',
                            'confidence': 0.0,
                            'validation_score': 0.0,
                            'idp_field': 'unmatched_monetary',
                            'field_type': 'monetary'
                        }
                        results.append(result)
                    
                    for code in unmatched.get('unmatched_codes', [])[:5]:  # Limit to 5
                        result = {
                            'filename': filename,
                            'search_term': 'UNMATCHED Code',
                            'page': 'Multiple',
                            'context': code.get('context', ''),
                            'match': code['value'],
                            'extraction_method': 'IDP-Unmatched',
                            'confidence': 0.0,
                            'validation_score': 0.0,
                            'idp_field': 'unmatched_code',
                            'field_type': 'code'
                        }
                        results.append(result)
                    
                    # Store full IDP results for Excel export
                    results.append({
                        'filename': filename,
                        'idp_full_results': idp_results,
                        'is_idp_metadata': True
                    })
                    
                    # Status update
                    found_count = sum(1 for field in idp_results['required_fields'].values() if field['best_match'])
                    total_fields = len(idp_results['required_fields'])
                    quality = idp_results['quality_metrics']
                    
                    self.root.after(0, lambda f=found_count, t=total_fields, q=quality, fn=filename: 
                                   self.results_text.insert(tk.END, 
                                   f"🎯 IDP: {fn} - Found {f}/{t} fields (Quality: {q['success_rate']:.1f}%)\n"))
                
                elif self.insurance_mode_var.get():
                    # Insurance Mode - use specialized extraction
                    insurance_data = self.insurance_extractor.extract_insurance_data(full_text, filename)
                    
                    # Convert insurance data to results format
                    for field_key, field_info in insurance_data.items():
                        if field_info['found']:
                            result = {
                                'filename': filename,
                                'search_term': field_info['field_name'],
                                'page': 'Multiple',  # Insurance fields can be on any page
                                'context': f"Found: {field_info['value']}",
                                'match': field_info['value'],
                                'extraction_method': extraction_method,
                                'insurance_field': field_key
                            }
                            results.append(result)
                    
                    # Add summary info
                    found_count = sum(1 for field in insurance_data.values() if field['found'])
                    total_fields = len(insurance_data)
                    
                    self.root.after(0, lambda f=found_count, t=total_fields, fn=filename: 
                                   self.results_text.insert(tk.END, 
                                   f"🏢 {fn}: Found {f}/{t} insurance fields\n"))
                
                else:
                    # Normal Mode - search for each term
                    for search_term in self.search_terms:
                        matches = self.find_matches(full_text, search_term, page_texts)
                        
                        for match in matches:
                            result = {
                                'filename': filename,
                                'filepath': file_path,
                                'search_term': search_term,
                                'page_number': match.get('page', 'Unknown'),
                                'match_text': match.get('match', ''),
                                'context': match.get('context', ''),
                                'extraction_method': extraction_method,
                                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }
                            results.append(result)
        
        except Exception as e:
            self.logger.error(f"Error extracting from {file_path}: {str(e)}")
            raise
        
        return results
    
    def extract_text_with_ocr(self, pdf, filename: str) -> Dict[int, str]:
        """Extract text from PDF using OCR"""
        page_texts = {}
        
        try:
            for page_num, page in enumerate(pdf.pages, 1):
                try:
                    # Update status for current page
                    self.root.after(0, lambda p=page_num, total=len(pdf.pages): 
                                   self.status_label.config(text=f"OCR processing {filename} - Page {p}/{total}"))
                    
                    # Convert page to image
                    page_image = page.to_image(resolution=300)  # High resolution for better OCR
                    pil_image = page_image.original
                    
                    # Perform OCR on the image
                    ocr_text = pytesseract.image_to_string(pil_image, config='--psm 6')
                    
                    if ocr_text.strip():
                        page_texts[page_num] = ocr_text
                        self.logger.info(f"OCR extracted {len(ocr_text)} characters from page {page_num} of {filename}")
                    else:
                        page_texts[page_num] = ""
                        self.logger.warning(f"No OCR text found on page {page_num} of {filename}")
                        
                except Exception as e:
                    self.logger.error(f"OCR failed for page {page_num} of {filename}: {str(e)}")
                    page_texts[page_num] = ""
                    
        except Exception as e:
            self.logger.error(f"OCR processing failed for {filename}: {str(e)}")
            # Return empty dict if OCR completely fails
            return {}
        
        return page_texts
    
    def find_matches(self, text: str, search_term: str, page_texts: Dict) -> List[Dict]:
        """Find all matches for a search term in the text"""
        matches = []
        
        try:
            # Prepare search pattern
            if self.regex_var.get():
                pattern = search_term
            else:
                # Escape regex special characters
                pattern = re.escape(search_term)
                if self.whole_words_var.get():
                    pattern = r'\b' + pattern + r'\b'
            
            # Set case sensitivity
            flags = 0 if self.case_sensitive_var.get() else re.IGNORECASE
            
            # Search in each page
            for page_num, page_text in page_texts.items():
                for match in re.finditer(pattern, page_text, flags):
                    match_text = match.group()
                    start_pos = match.start()
                    end_pos = match.end()
                    
                    # Get context if requested
                    context = ""
                    if self.context_var.get():
                        context_length = int(self.context_length.get() or 100)
                        context_start = max(0, start_pos - context_length)
                        context_end = min(len(page_text), end_pos + context_length)
                        context = page_text[context_start:context_end]
                        # Clean up context
                        context = ' '.join(context.split())
                    
                    matches.append({
                        'page': page_num,
                        'match': match_text,
                        'context': context,
                        'position': start_pos
                    })
        
        except re.error as e:
            self.logger.error(f"Regex error for pattern '{search_term}': {str(e)}")
            # Fallback to simple string search
            if search_term in text:
                matches.append({
                    'page': 'Unknown',
                    'match': search_term,
                    'context': '',
                    'position': 0
                })
        
        return matches
    
    def processing_complete(self):
        """Called when processing is complete"""
        self.is_processing = False
        
        # Display results
        if self.extracted_data:
            self.display_results()
            self.export_button.config(state=tk.NORMAL)
            self.status_label.config(text=f"Processing complete! Found {len(self.extracted_data)} matches.")
        else:
            self.results_text.insert(tk.END, "No matches found in the selected files.\n")
            self.status_label.config(text="Processing complete - No matches found.")
        
        self.reset_ui()
    
    def display_results(self):
        """Display extraction results in the text widget"""
        self.results_text.delete(1.0, tk.END)
        
        # Group results by file and search term
        by_file = {}
        for result in self.extracted_data:
            filename = result['filename']
            if filename not in by_file:
                by_file[filename] = {}
            
            term = result['search_term']
            if term not in by_file[filename]:
                by_file[filename][term] = []
            
            by_file[filename][term].append(result)
        
        # Display grouped results
        for filename, terms_data in by_file.items():
            self.results_text.insert(tk.END, f"\n=== {filename} ===\n")
            
            for term, matches in terms_data.items():
                self.results_text.insert(tk.END, f"\n  Search term: '{term}' ({len(matches)} matches)\n")
                
                for i, match in enumerate(matches, 1):
                    page = match.get('page', 'Unknown')
                    match_text = match.get('match', '')
                    context = match.get('context', '')
                    extraction_method = match.get('extraction_method', 'Normal')
                    
                    method_indicator = " [OCR]" if extraction_method == "OCR" else ""
                    self.results_text.insert(tk.END, f"    {i}. Page {page}{method_indicator}: {match_text}\n")
                    
                    if context and len(context) > len(match_text):
                        # Show context with match highlighted
                        self.results_text.insert(tk.END, f"       Context: ...{context}...\n")
                    
                    self.results_text.insert(tk.END, "\n")
        
        # Summary
        total_matches = len(self.extracted_data)
        unique_files = len(set(result['filename'] for result in self.extracted_data))
        unique_terms = len(set(result['search_term'] for result in self.extracted_data))
        
        self.results_text.insert(tk.END, f"\n=== SUMMARY ===\n")
        self.results_text.insert(tk.END, f"Total matches: {total_matches}\n")
        self.results_text.insert(tk.END, f"Files with matches: {unique_files}\n")
        self.results_text.insert(tk.END, f"Search terms found: {unique_terms}\n")
    
    def export_to_excel(self):
        """Export extracted data to Excel file"""
        if not self.extracted_data:
            messagebox.showwarning("Warning", "No data to export.")
            return
        
        # Ask for save location
        filename = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            # Check the mode we're in
            if self.idp_mode_var.get():
                self.export_idp_data(filename)
            elif self.insurance_mode_var.get():
                self.export_insurance_data(filename)
            else:
                self.export_normal_data(filename)
                
        except Exception as e:
            self.logger.error(f"Export error: {str(e)}")
            messagebox.showerror("Error", f"Failed to export data:\n{str(e)}")
    
    def export_idp_data(self, filename: str):
        """Export IDP comprehensive data with unmatched data visibility"""
        try:
            # Find IDP metadata
            idp_metadata = None
            for result in self.extracted_data:
                if result.get('is_idp_metadata'):
                    idp_metadata = result['idp_full_results']
                    break
            
            if idp_metadata:
                # Use the IDP extractor's comprehensive Excel creator
                success = self.idp_extractor.create_comprehensive_excel(idp_metadata, filename)
                
                if success:
                    messagebox.showinfo("Success", f"IDP comprehensive data exported successfully to:\n{filename}\n\nIncludes:\n• Required Fields (15 fields)\n• All Candidates (100% visibility)\n• Unmatched Data (for manual review)\n• Quality Metrics\n• Processing Log")
                    self.status_label.config(text=f"IDP data exported: {os.path.basename(filename)}")
                else:
                    raise Exception("IDP export failed")
            else:
                # Fallback to regular export with enhanced data
                self.export_idp_fallback(filename)
                
        except Exception as e:
            self.logger.error(f"IDP export error: {str(e)}")
            raise
    
    def export_insurance_data(self, filename: str):
        """Export insurance-specific data using the specialized insurance extractor"""
        # Transform extracted data into insurance format
        insurance_data = []
        
        # Group by filename
        by_file = {}
        for result in self.extracted_data:
            file_key = result['filename']
            if file_key not in by_file:
                by_file[file_key] = {
                    'filename': file_key,
                    'extraction_method': result.get('extraction_method', 'Normal'),
                    'insurance_data': {}
                }
            
            # Add insurance field data
            if 'insurance_field' in result:
                field_key = result['insurance_field']
                by_file[file_key]['insurance_data'][field_key] = {
                    'field_name': result['search_term'],
                    'value': result.get('match', result.get('context', 'Found')),
                    'found': True
                }
        
        # Convert to list format expected by insurance extractor
        for file_data in by_file.values():
            insurance_data.append(file_data)
        
        # Use the insurance extractor's specialized Excel creator
        success = self.insurance_extractor.create_insurance_excel(insurance_data, filename)
        
        if success:
            messagebox.showinfo("Success", f"Insurance data exported successfully to:\n{filename}")
            self.status_label.config(text=f"Insurance data exported to Excel: {os.path.basename(filename)}")
        else:
            raise Exception("Insurance export failed")
    
    def export_normal_data(self, filename: str):
        """Export normal search data using the standard format"""
        try:
            # Create DataFrame
            df = pd.DataFrame(self.extracted_data)
            
            # Create Excel file with formatting
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Extracted Data', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Extracted Data']
                
                # Format headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:  # First row (headers)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    
                    # Set width with some padding, but cap it
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add summary sheet
                summary_data = self.create_summary_data()
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format summary sheet
                summary_sheet = writer.sheets['Summary']
                for cell in summary_sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
            
            messagebox.showinfo("Success", f"Data exported successfully to:\n{filename}")
            self.status_label.config(text=f"Data exported to Excel: {os.path.basename(filename)}")
            
        except Exception as e:
            self.logger.error(f"Export error: {str(e)}")
            messagebox.showerror("Error", f"Failed to export data:\n{str(e)}")
    
    def create_summary_data(self) -> List[Dict]:
        """Create summary data for the Excel export"""
        summary = []
        
        # Count by file
        file_counts = {}
        for result in self.extracted_data:
            filename = result['filename']
            file_counts[filename] = file_counts.get(filename, 0) + 1
        
        for filename, count in file_counts.items():
            summary.append({
                'Type': 'File',
                'Name': filename,
                'Matches': count
            })
        
        # Count by search term
        term_counts = {}
        for result in self.extracted_data:
            term = result['search_term']
            term_counts[term] = term_counts.get(term, 0) + 1
        
        for term, count in term_counts.items():
            summary.append({
                'Type': 'Search Term',
                'Name': term,
                'Matches': count
            })
        
        return summary
    
    def reset_ui(self):
        """Reset UI elements after processing"""
        self.process_button.config(state=tk.NORMAL)
        self.progress_var.set(0)

    def export_idp_fallback(self, filename: str):
        """Fallback export for IDP data when metadata is not available"""
        try:
            # Filter out metadata entries
            filtered_data = [result for result in self.extracted_data if not result.get('is_idp_metadata')]
            
            # Create enhanced DataFrame with IDP-specific columns
            df_data = []
            for result in filtered_data:
                row = {
                    'Filename': result['filename'],
                    'Field Name': result['search_term'],
                    'Value': result['match'],
                    'Extraction Method': result.get('extraction_method', 'Unknown'),
                    'Confidence': f"{result.get('confidence', 0):.2f}",
                    'Validation Score': f"{result.get('validation_score', 0):.2f}",
                    'Field Type': result.get('field_type', 'unknown'),
                    'Context': result.get('context', '')[:200] + '...' if len(result.get('context', '')) > 200 else result.get('context', ''),
                    'Status': 'Found' if result.get('confidence', 0) > 0 else 'Unmatched'
                }
                df_data.append(row)
            
            df = pd.DataFrame(df_data)
            
            # Create Excel with formatting
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='IDP Extraction Results', index=False)
                
                # Format the sheet
                from openpyxl.styles import Font, PatternFill, Alignment
                ws = writer.sheets['IDP Extraction Results']
                
                # Header formatting
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            
            messagebox.showinfo("Success", f"IDP data exported (fallback mode) to:\n{filename}")
            self.status_label.config(text=f"IDP data exported (fallback): {os.path.basename(filename)}")
            
        except Exception as e:
            self.logger.error(f"IDP fallback export error: {str(e)}")
            raise

def main():
    """Main application entry point"""
    root = tk.Tk()
    
    # Set app icon and style
    try:
        # Configure ttk style
        style = ttk.Style()
        style.theme_use('clam')  # Use a modern theme
        
        # Configure custom style for accent button
        style.configure('Accent.TButton', foreground='white')
        style.map('Accent.TButton', background=[('active', '#0078d4'), ('!active', '#106ebe')])
        
    except Exception as e:
        print(f"Style configuration error: {e}")
    
    app = PDFDataExtractor(root)
    
    # Center window on screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()

if __name__ == "__main__":
    main() 
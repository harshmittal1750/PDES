#!/usr/bin/env python3
"""
Insurance PDF Data Extractor - Enhanced Mode
Automatically extracts specific insurance policy fields from PDFs with high accuracy
"""

import re
import logging
from typing import Dict, List, Optional, Tuple, Set
from datetime import datetime
import pandas as pd
import difflib
from dataclasses import dataclass

@dataclass
class ExtractionResult:
    """Container for extraction results with confidence scoring"""
    value: str
    confidence: float
    method: str
    position: int
    context: str

class EnhancedInsuranceExtractor:
    """Enhanced extractor for insurance policy documents with high accuracy"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # Define the fields we're looking for with aliases
        self.insurance_fields = {
            'policy_no': {
                'name': 'Policy no.',
                'aliases': ['policy number', 'policy no', 'certificate no', 'certificate number', 
                           'policy ref', 'policy reference', 'cert no', 'cert number']
            },
            'insured_name': {
                'name': 'Insured name',
                'aliases': ['insured name', 'name of insured', 'policy holder', 'insured', 
                           'policyholder name', 'customer name', 'assured name']
            },
            'insurer_name': {
                'name': 'Insurer name',
                'aliases': ['insurer name', 'insurance company', 'company name', 'insurer', 
                           'company', 'underwriter', 'carrier']
            },
            'engine_no': {
                'name': 'Engine no.',
                'aliases': ['engine no', 'engine number', 'engine', 'engine serial', 
                           'motor no', 'motor number']
            },
            'chassis_no': {
                'name': 'Chassis no.',
                'aliases': ['chassis no', 'chassis number', 'vin', 'chassis', 'vehicle identification number',
                           'frame no', 'frame number', 'vehicle id']
            },
            'cheque_no': {
                'name': 'Cheque no.',
                'aliases': ['cheque no', 'check no', 'cheque number', 'check number', 'cheque',
                           'check', 'payment ref', 'payment reference']
            },
            'cheque_date': {
                'name': 'Cheque date',
                'aliases': ['cheque date', 'check date', 'payment date', 'date of payment',
                           'payment on', 'paid on', 'transaction date']
            },
            'bank_name': {
                'name': 'Bank name',
                'aliases': ['bank name', 'bank', 'drawn on', 'issuing bank', 'paying bank',
                           'financial institution']
            },
            'net_od_premium': {
                'name': 'Net own damage premium amount',
                'aliases': ['net od premium', 'own damage premium', 'od premium', 'net own damage',
                           'comprehensive premium', 'property damage premium', 'vehicle premium']
            },
            'net_liability_premium': {
                'name': 'Net liability premium amount',
                'aliases': ['net liability premium', 'liability premium', 'tp premium', 'third party premium',
                           'liability amount', 'tp amount', 'third party amount']
            },
            'total_premium': {
                'name': 'Total premium amount',
                'aliases': ['total premium', 'net premium', 'premium amount', 'base premium',
                           'subtotal', 'premium subtotal', 'total amount']
            },
            'gst_amount': {
                'name': 'GST amount',
                'aliases': ['gst', 'service tax', 'tax amount', 'igst', 'cgst', 'sgst',
                           'tax', 'vat', 'sales tax']
            },
            'gross_premium': {
                'name': 'Gross premium paid',
                'aliases': ['gross premium', 'total amount', 'amount paid', 'final amount',
                           'total payable', 'grand total', 'amount due']
            },
            'car_model': {
                'name': 'Car model',
                'aliases': ['model', 'vehicle model', 'make model', 'car model', 'vehicle make',
                           'make and model', 'vehicle description']
            },
            'body_type': {
                'name': 'Body type',
                'aliases': ['body type', 'vehicle type', 'type of vehicle', 'category',
                           'vehicle category', 'classification']
            }
        }
        
        # Create enhanced search patterns
        self.search_patterns = self.create_enhanced_patterns()
        self.contextual_patterns = self.create_contextual_patterns()
        self.table_patterns = self.create_table_patterns()
    
    def create_enhanced_patterns(self) -> Dict[str, List[str]]:
        """Create comprehensive search patterns with multiple variations"""
        patterns = {
            'policy_no': [
                # Standard patterns
                r'(?:policy|certificate|cert)\s*(?:no|number|ref|reference)\.?\s*:?\s*([A-Z0-9\-/]{4,})',
                r'policy\s*:?\s*([A-Z0-9\-/]{6,})',
                r'cert\.?\s*no\.?\s*:?\s*([A-Z0-9\-/]{6,})',
                # Table patterns
                r'([A-Z0-9\-/]{8,})\s*(?:\n|\s{3,})', # Long codes often standalone
                # Context-aware patterns
                r'(?:policy|certificate)\s+(?:is\s+)?([A-Z0-9\-/]{6,})',
            ],
            
            'insured_name': [
                r'(?:insured|policy\s*holder|customer|assured)\s*(?:name)?\s*:?\s*([A-Za-z\s\.\,]{3,50})(?:\n|$|[0-9])',
                r'name\s*of\s*(?:insured|policy\s*holder)\s*:?\s*([A-Za-z\s\.\,]{3,50})(?:\n|$)',
                r'(?:mr|mrs|ms|dr|m/s)\.?\s+([A-Za-z\s\.\,]{3,50})(?:\n|$)',
                # Table patterns
                r'^([A-Z][A-Za-z\s\.\,]{10,40})$', # Names in table cells
            ],
            
            'insurer_name': [
                r'(?:insurer|insurance\s*company|company|underwriter|carrier)\s*(?:name)?\s*:?\s*([A-Za-z\s&\.\,\-]{3,})(?:\n|$)',
                r'(?:insured\s*with|covered\s*by|policy\s*by)\s*:?\s*([A-Za-z\s&\.\,\-]{3,})(?:\n|$)',
                r'([A-Za-z\s&\.\-]{3,})\s*(?:insurance|assurance|general\s*insurance)(?:\s|$)',
            ],
            
            'engine_no': [
                r'(?:engine|motor)\s*(?:no|number|serial|#)\.?\s*:?\s*([A-Z0-9]{4,})',
                r'engine\s*:?\s*([A-Z0-9]{6,})',
                r'e\.?\s*no\.?\s*:?\s*([A-Z0-9]{6,})',
            ],
            
            'chassis_no': [
                r'(?:chassis|vin|frame)\s*(?:no|number|#)\.?\s*:?\s*([A-Z0-9]{4,})',
                r'vehicle\s*identification\s*(?:no|number)\s*:?\s*([A-Z0-9]{17})', # VIN is 17 chars
                r'chassis\s*:?\s*([A-Z0-9]{6,})',
                r'c\.?\s*no\.?\s*:?\s*([A-Z0-9]{6,})',
            ],
            
            'cheque_no': [
                r'(?:cheque|check)\s*(?:no|number|#)\.?\s*:?\s*([0-9]{4,})',
                r'(?:payment|transaction)\s*(?:ref|reference)\s*:?\s*([0-9A-Z\-]{4,})',
                r'cheque\s*:?\s*([0-9]{6,})',
            ],
            
            'cheque_date': [
                r'(?:cheque|check|payment|transaction)\s*date\s*:?\s*([0-9]{1,2}[\/\-\.][0-9]{1,2}[\/\-\.][0-9]{2,4})',
                r'(?:paid|payment)\s*(?:on|date)\s*:?\s*([0-9]{1,2}[\/\-\.][0-9]{1,2}[\/\-\.][0-9]{2,4})',
                r'date\s*of\s*payment\s*:?\s*([0-9]{1,2}[\/\-\.][0-9]{1,2}[\/\-\.][0-9]{2,4})',
            ],
            
            'bank_name': [
                r'(?:bank|drawn\s*on|issuing\s*bank)\s*(?:name)?\s*:?\s*([A-Za-z\s&\.\,\-]{3,})(?:\n|$|branch)',
                r'([A-Za-z\s&\.\-]{3,})\s*bank(?:\s|$)',
                r'financial\s*institution\s*:?\s*([A-Za-z\s&\.\,\-]{3,})(?:\n|$)',
            ],
            
            # Enhanced monetary patterns with multiple currency formats
            'net_od_premium': [
                r'(?:net\s*)?(?:own\s*damage|od|comprehensive)\s*premium\s*(?:amount)?\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
                r'(?:property|vehicle)\s*(?:damage\s*)?premium\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
                r'od\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
            ],
            
            'net_liability_premium': [
                r'(?:net\s*)?(?:liability|tp|third\s*party)\s*premium\s*(?:amount)?\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
                r'third\s*party\s*(?:premium|amount)\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
                r'tp\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
            ],
            
            'total_premium': [
                r'(?:total|net|base)\s*premium\s*(?:amount)?\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
                r'premium\s*(?:subtotal|amount)\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
                r'subtotal\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
            ],
            
            'gst_amount': [
                r'(?:gst|igst|cgst|sgst|service\s*tax|tax)\s*(?:amount)?\s*(?:\(@[0-9]+%\))?\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
                r'tax\s*(?:component|amount)\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
                r'(?:total\s*)?tax\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
            ],
            
            'gross_premium': [
                r'(?:gross|total|final)\s*(?:premium|amount)\s*(?:paid|payable)?\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
                r'(?:grand\s*total|amount\s*payable|total\s*amount)\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
                r'amount\s*(?:paid|due)\s*:?\s*(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)',
            ],
            
            'car_model': [
                r'(?:make|model|vehicle\s*model|car\s*model)\s*:?\s*([A-Za-z0-9\s\-\/]{3,})(?:\n|$|year)',
                r'(?:make\s*[&\/]\s*model|vehicle\s*description)\s*:?\s*([A-Za-z0-9\s\-\/]{3,})(?:\n|$)',
                r'^([A-Z][A-Za-z0-9\s\-\/]{3,})(?:\s*[0-9]{4}|\n|$)', # Model with optional year
            ],
            
            'body_type': [
                r'(?:body\s*type|vehicle\s*type|category|classification)\s*:?\s*([A-Za-z\s\-]{3,})(?:\n|$)',
                r'type\s*of\s*vehicle\s*:?\s*([A-Za-z\s\-]{3,})(?:\n|$)',
                r'(?:sedan|hatchback|suv|coupe|convertible|wagon|truck|motorcycle|scooter)',
            ],
        }
        
        return patterns
    
    def create_contextual_patterns(self) -> Dict[str, List[str]]:
        """Create patterns that look for values near field labels"""
        return {
            'table_row': r'([^|\n]+)\s*\|\s*([^|\n]+)', # Basic table detection
            'colon_separated': r'([^:\n]+)\s*:\s*([^:\n]+)', # Label: Value
            'nearby_value': r'({label})\s*[:\-]?\s*([A-Za-z0-9\s\.\-\/,]+)(?:\n|$|[|])',
        }
    
    def create_table_patterns(self) -> Dict[str, List[str]]:
        """Patterns for extracting data from table structures"""
        return {
            'horizontal_table': r'([^|\t\n]+)[\|\t]+([^|\t\n]+)', # Pipe or tab separated
            'vertical_spacing': r'([A-Za-z\s:]+)\s{3,}([A-Za-z0-9\s\-/,\.]+)', # Space separated columns
            'underline_headers': r'([A-Za-z\s:]+)\n[\-=_\s]+\n([A-Za-z0-9\s\-/,\.]+)', # Headers with underlines
        }
    
    def extract_insurance_data(self, text: str, filename: str) -> Dict[str, Dict]:
        """Enhanced extraction with multiple strategies and confidence scoring"""
        results = {}
        text_lower = text.lower()
        
        self.logger.info(f"Starting enhanced extraction from {filename}")
        
        for field_key, field_info in self.insurance_fields.items():
            field_name = field_info['name']
            
            # Multi-pass extraction strategy
            extraction_results = []
            
            # Pass 1: Direct pattern matching
            direct_result = self.extract_with_patterns(text_lower, field_key)
            if direct_result:
                extraction_results.append(direct_result)
            
            # Pass 2: Contextual extraction (look near field labels)
            contextual_result = self.extract_contextually(text, text_lower, field_key, field_info['aliases'])
            if contextual_result:
                extraction_results.append(contextual_result)
            
            # Pass 3: Table extraction
            table_result = self.extract_from_tables(text, text_lower, field_key, field_info['aliases'])
            if table_result:
                extraction_results.append(table_result)
            
            # Pass 4: Fuzzy matching for labels
            fuzzy_result = self.extract_with_fuzzy_matching(text, text_lower, field_key, field_info['aliases'])
            if fuzzy_result:
                extraction_results.append(fuzzy_result)
            
            # Select best result based on confidence
            best_result = self.select_best_result(extraction_results)
            
            if best_result:
                results[field_key] = {
                    'field_name': field_name,
                    'value': best_result.value,
                    'confidence': best_result.confidence,
                    'method': best_result.method,
                    'context': best_result.context[:200] + '...' if len(best_result.context) > 200 else best_result.context,
                    'found': True
                }
                self.logger.info(f"Found {field_name}: {best_result.value} (confidence: {best_result.confidence:.2f}, method: {best_result.method})")
            else:
                results[field_key] = {
                    'field_name': field_name,
                    'value': 'Not found',
                    'confidence': 0.0,
                    'method': 'None',
                    'context': '',
                    'found': False
                }
                self.logger.warning(f"Could not find {field_name} in {filename}")
        
        return results
    
    def extract_with_patterns(self, text: str, field_key: str) -> Optional[ExtractionResult]:
        """Extract using direct regex patterns"""
        patterns = self.search_patterns.get(field_key, [])
        
        for i, pattern in enumerate(patterns):
            matches = list(re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE))
            for match in matches:
                value = match.group(1).strip() if match.groups() else match.group().strip()
                if value and len(value) > 1:
                    cleaned_value = self.clean_value(value, field_key)
                    if cleaned_value and self.validate_field(cleaned_value, field_key):
                        confidence = 0.9 - (i * 0.1) # Higher confidence for earlier patterns
                        context = self.get_context(text, match.start(), match.end())
                        return ExtractionResult(
                            value=cleaned_value,
                            confidence=confidence,
                            method=f"Direct Pattern {i+1}",
                            position=match.start(),
                            context=context
                        )
        return None
    
    def extract_contextually(self, original_text: str, text_lower: str, field_key: str, aliases: List[str]) -> Optional[ExtractionResult]:
        """Extract by looking for values near field labels"""
        for alias in aliases:
            # Look for the alias followed by a separator and then the value
            patterns = [
                rf'{re.escape(alias.lower())}\s*:?\s*([A-Za-z0-9\s\.\-\/,]+?)(?:\n|$|[|])',
                rf'{re.escape(alias.lower())}\s*[:\-]\s*([A-Za-z0-9\s\.\-\/,]+?)(?:\n|$|[|])',
                rf'{re.escape(alias.lower())}\s+([A-Za-z0-9\s\.\-\/,]+?)(?:\n|$|[|])',
            ]
            
            for pattern in patterns:
                matches = list(re.finditer(pattern, text_lower, re.IGNORECASE | re.MULTILINE))
                for match in matches:
                    # Get the original case value from the original text
                    start, end = match.span(1)
                    value = original_text[start:end].strip()
                    
                    if value and len(value) > 1:
                        cleaned_value = self.clean_value(value, field_key)
                        if cleaned_value and self.validate_field(cleaned_value, field_key):
                            confidence = 0.85
                            context = self.get_context(original_text, match.start(), match.end())
                            return ExtractionResult(
                                value=cleaned_value,
                                confidence=confidence,
                                method=f"Contextual ({alias})",
                                position=match.start(),
                                context=context
                            )
        return None
    
    def extract_from_tables(self, original_text: str, text_lower: str, field_key: str, aliases: List[str]) -> Optional[ExtractionResult]:
        """Extract from table-like structures"""
        # Split text into lines and look for tabular patterns
        lines = original_text.split('\n')
        
        for i, line in enumerate(lines):
            line_lower = line.lower()
            
            # Check if this line contains a field label
            for alias in aliases:
                if alias.lower() in line_lower:
                    # Look for value in the same line or next few lines
                    for j in range(max(0, i-1), min(len(lines), i+3)):
                        potential_line = lines[j]
                        
                        # Try different table extraction patterns
                        table_patterns = [
                            r'([^|\t]+)[\|\t]+([^|\t]+)', # Pipe or tab separated
                            r'([^:]+):\s*([^:]+)', # Colon separated
                            r'(.+?)\s{3,}(.+)', # Multiple spaces
                        ]
                        
                        for pattern in table_patterns:
                            matches = re.finditer(pattern, potential_line)
                            for match in matches:
                                # Check both parts of the match
                                part1, part2 = match.groups()
                                
                                # If part1 contains the alias, part2 is likely the value
                                if alias.lower() in part1.lower():
                                    value = part2.strip()
                                elif alias.lower() in part2.lower():
                                    value = part1.strip()
                                else:
                                    continue
                                
                                if value and len(value) > 1:
                                    cleaned_value = self.clean_value(value, field_key)
                                    if cleaned_value and self.validate_field(cleaned_value, field_key):
                                        confidence = 0.75
                                        context = '\n'.join(lines[max(0, i-1):min(len(lines), i+3)])
                                        return ExtractionResult(
                                            value=cleaned_value,
                                            confidence=confidence,
                                            method=f"Table Extraction ({alias})",
                                            position=0,
                                            context=context
                                        )
        return None
    
    def extract_with_fuzzy_matching(self, original_text: str, text_lower: str, field_key: str, aliases: List[str]) -> Optional[ExtractionResult]:
        """Extract using fuzzy string matching for field labels"""
        lines = original_text.split('\n')
        
        for i, line in enumerate(lines):
            words = re.findall(r'\b\w+\b', line.lower())
            line_text = ' '.join(words)
            
            # Check for fuzzy matches with aliases
            for alias in aliases:
                alias_words = alias.lower().split()
                
                # Use fuzzy matching to find similar phrases
                similarity = difflib.SequenceMatcher(None, alias.lower(), line_text).ratio()
                
                if similarity > 0.7:  # 70% similarity threshold
                    # Look for values in surrounding lines
                    search_lines = lines[max(0, i-1):min(len(lines), i+3)]
                    
                    for search_line in search_lines:
                        # Extract potential values using field-specific patterns
                        if field_key in ['net_od_premium', 'net_liability_premium', 'total_premium', 'gst_amount', 'gross_premium']:
                            # Look for monetary values
                            money_matches = re.finditer(r'(?:rs\.?|₹|inr)?\s*([0-9,]+(?:\.[0-9]{2})?)', search_line, re.IGNORECASE)
                            for match in money_matches:
                                value = match.group(1)
                                cleaned_value = self.clean_value(value, field_key)
                                if cleaned_value and self.validate_field(cleaned_value, field_key):
                                    confidence = 0.6 + (similarity * 0.2)
                                    context = '\n'.join(search_lines)
                                    return ExtractionResult(
                                        value=cleaned_value,
                                        confidence=confidence,
                                        method=f"Fuzzy Match ({alias})",
                                        position=0,
                                        context=context
                                    )
                        elif field_key in ['policy_no', 'engine_no', 'chassis_no', 'cheque_no']:
                            # Look for alphanumeric codes
                            code_matches = re.finditer(r'\b([A-Z0-9\-/]{4,})\b', search_line)
                            for match in code_matches:
                                value = match.group(1)
                                cleaned_value = self.clean_value(value, field_key)
                                if cleaned_value and self.validate_field(cleaned_value, field_key):
                                    confidence = 0.6 + (similarity * 0.2)
                                    context = '\n'.join(search_lines)
                                    return ExtractionResult(
                                        value=cleaned_value,
                                        confidence=confidence,
                                        method=f"Fuzzy Match ({alias})",
                                        position=0,
                                        context=context
                                    )
        return None
    
    def validate_field(self, value: str, field_key: str) -> bool:
        """Validate extracted values based on field type"""
        if not value or len(value.strip()) < 1:
            return False
        
        value = value.strip()
        
        # Field-specific validation
        if field_key in ['net_od_premium', 'net_liability_premium', 'total_premium', 'gst_amount', 'gross_premium']:
            # Monetary fields - should be numeric
            return bool(re.match(r'^[0-9,]+(?:\.[0-9]{1,2})?$', value))
        
        elif field_key == 'policy_no':
            # Policy numbers - alphanumeric, reasonable length
            return len(value) >= 4 and len(value) <= 30 and bool(re.match(r'^[A-Z0-9\-/]+$', value))
        
        elif field_key in ['engine_no', 'chassis_no']:
            # Engine/Chassis - alphanumeric codes
            return len(value) >= 4 and len(value) <= 25 and bool(re.match(r'^[A-Z0-9]+$', value))
        
        elif field_key == 'cheque_no':
            # Cheque numbers - numeric
            return len(value) >= 4 and len(value) <= 15 and bool(re.match(r'^[0-9]+$', value))
        
        elif field_key == 'cheque_date':
            # Dates - valid date format
            return bool(re.match(r'^[0-9]{1,2}[\/\-\.][0-9]{1,2}[\/\-\.][0-9]{2,4}$', value))
        
        elif field_key in ['insured_name', 'insurer_name', 'bank_name']:
            # Names - should contain alphabetic characters
            return len(value) >= 3 and bool(re.search(r'[A-Za-z]', value))
        
        elif field_key == 'car_model':
            # Car models - alphanumeric with spaces
            return len(value) >= 2 and bool(re.match(r'^[A-Za-z0-9\s\-\/]+$', value))
        
        elif field_key == 'body_type':
            # Body types - common vehicle types
            common_types = ['sedan', 'hatchback', 'suv', 'coupe', 'convertible', 'wagon', 'truck', 'motorcycle', 'scooter', 'van']
            return len(value) >= 3 and (value.lower() in common_types or bool(re.match(r'^[A-Za-z\s\-]+$', value)))
        
        return True  # Default validation
    
    def get_context(self, text: str, start: int, end: int, context_length: int = 100) -> str:
        """Extract context around a match"""
        context_start = max(0, start - context_length)
        context_end = min(len(text), end + context_length)
        context = text[context_start:context_end]
        return ' '.join(context.split())  # Clean whitespace
    
    def select_best_result(self, results: List[ExtractionResult]) -> Optional[ExtractionResult]:
        """Select the best extraction result based on confidence and validation"""
        if not results:
            return None
        
        # Sort by confidence score
        results.sort(key=lambda x: x.confidence, reverse=True)
        
        # Return the highest confidence result
        return results[0]
    
    def clean_value(self, value: str, field_key: str) -> str:
        """Enhanced cleaning and validation of extracted values"""
        if not value:
            return ''
        
        # Remove extra whitespace and normalize
        value = re.sub(r'\s+', ' ', value.strip())
        
        # Field-specific cleaning
        if field_key in ['net_od_premium', 'net_liability_premium', 'total_premium', 'gst_amount', 'gross_premium']:
            # Clean monetary values - remove currency symbols and extra characters
            value = re.sub(r'[^\d,\.]', '', value)
            value = value.replace(',', '')  # Remove thousand separators
            # Ensure proper decimal format
            if '.' in value:
                parts = value.split('.')
                if len(parts) == 2 and len(parts[1]) <= 2:
                    value = f"{parts[0]}.{parts[1]}"
                else:
                    value = parts[0]  # Take integer part if decimal is malformed
            
        elif field_key == 'cheque_date':
            # Standardize date format and validate
            value = re.sub(r'[^\d\/\-\.]', '', value)
            # Convert different separators to forward slash
            value = re.sub(r'[\-\.]', '/', value)
            
        elif field_key in ['policy_no', 'engine_no', 'chassis_no']:
            # Clean alphanumeric codes - remove non-alphanumeric except hyphens and forward slashes
            value = re.sub(r'[^\w\-/]', '', value).upper()
            
        elif field_key == 'cheque_no':
            # Clean cheque numbers - numbers only
            value = re.sub(r'[^\d]', '', value)
        
        elif field_key in ['insured_name', 'insurer_name', 'bank_name', 'car_model']:
            # Clean names and text fields - remove excessive punctuation
            value = re.sub(r'[^\w\s&\.\-,]', '', value)
            value = re.sub(r'\s+', ' ', value)  # Normalize spaces
            value = value.title()  # Title case for readability
            
        elif field_key == 'body_type':
            # Clean body type - standardize common terms
            value = value.lower().strip()
            body_type_mappings = {
                'saloon': 'sedan',
                'estate': 'wagon',
                'sports utility vehicle': 'suv',
                'sport utility vehicle': 'suv',
                'multi purpose vehicle': 'van',
                'mpv': 'van'
            }
            value = body_type_mappings.get(value, value)
            value = value.title()
        
        return value.strip() if value else ''
    
    # Keep the existing Excel creation methods with enhancements
    def create_insurance_excel(self, extracted_data: List[Dict], output_path: str) -> bool:
        """Create an enhanced Excel file for insurance data with confidence scores"""
        try:
            excel_data = []
            
            for file_data in extracted_data:
                filename = file_data['filename']
                extraction_method = file_data.get('extraction_method', 'Unknown')
                fields = file_data['insurance_data']
                
                row = {
                    'Filename': filename,
                    'Extraction Method': extraction_method,
                    'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Overall Confidence': self.calculate_overall_confidence(fields)
                }
                
                # Add all insurance fields with confidence scores
                for field_key, field_info in fields.items():
                    row[field_info['field_name']] = field_info['value']
                    row[f"{field_info['field_name']} (Confidence)"] = f"{field_info.get('confidence', 0):.2f}"
                    row[f"{field_info['field_name']} (Method)"] = field_info.get('method', 'N/A')
                    row[f"{field_info['field_name']} (Found)"] = 'Yes' if field_info['found'] else 'No'
                
                excel_data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(excel_data)
            
            # Create Excel writer with multiple sheets
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Main data sheet
                df.to_excel(writer, sheet_name='Insurance Data', index=False)
                
                # Summary sheet with enhanced statistics
                summary_data = self.create_enhanced_summary_stats(extracted_data)
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Confidence analysis sheet
                confidence_data = self.create_confidence_analysis(extracted_data)
                confidence_df = pd.DataFrame(confidence_data)
                confidence_df.to_excel(writer, sheet_name='Confidence Analysis', index=False)
                
                # Format the sheets
                self.format_enhanced_excel_sheets(writer, df, summary_df, confidence_df)
            
            self.logger.info(f"Enhanced insurance Excel file created: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating enhanced insurance Excel file: {e}")
            return False
    
    def calculate_overall_confidence(self, fields: Dict) -> float:
        """Calculate overall confidence for a document"""
        confidences = [field.get('confidence', 0) for field in fields.values() if field.get('found')]
        return sum(confidences) / len(confidences) if confidences else 0.0
    
    def create_enhanced_summary_stats(self, extracted_data: List[Dict]) -> List[Dict]:
        """Create enhanced summary statistics including confidence metrics"""
        total_files = len(extracted_data)
        field_stats = {}
        
        # Initialize enhanced field statistics
        for field_key, field_info in self.insurance_fields.items():
            field_stats[field_key] = {
                'field_name': field_info['name'],
                'found_count': 0,
                'missing_count': 0,
                'avg_confidence': 0,
                'high_confidence_count': 0,  # confidence > 0.8
                'low_confidence_count': 0,   # confidence < 0.5
                'success_rate': 0
            }
        
        # Collect statistics
        for file_data in extracted_data:
            fields = file_data['insurance_data']
            for field_key, field_info in fields.items():
                if field_info['found']:
                    field_stats[field_key]['found_count'] += 1
                    confidence = field_info.get('confidence', 0)
                    field_stats[field_key]['avg_confidence'] += confidence
                    
                    if confidence > 0.8:
                        field_stats[field_key]['high_confidence_count'] += 1
                    elif confidence < 0.5:
                        field_stats[field_key]['low_confidence_count'] += 1
                else:
                    field_stats[field_key]['missing_count'] += 1
        
        # Calculate final statistics
        summary_rows = []
        for field_key, stats in field_stats.items():
            success_rate = (stats['found_count'] / total_files) * 100 if total_files > 0 else 0
            avg_confidence = stats['avg_confidence'] / stats['found_count'] if stats['found_count'] > 0 else 0
            
            summary_rows.append({
                'Field Name': stats['field_name'],
                'Found in Files': stats['found_count'],
                'Missing in Files': stats['missing_count'],
                'Success Rate (%)': round(success_rate, 1),
                'Avg Confidence': round(avg_confidence, 2),
                'High Confidence (>80%)': stats['high_confidence_count'],
                'Low Confidence (<50%)': stats['low_confidence_count'],
                'Total Files': total_files
            })
        
        return summary_rows
    
    def create_confidence_analysis(self, extracted_data: List[Dict]) -> List[Dict]:
        """Create confidence analysis data"""
        confidence_analysis = []
        
        for file_data in extracted_data:
            filename = file_data['filename']
            fields = file_data['insurance_data']
            overall_confidence = self.calculate_overall_confidence(fields)
            
            confidence_analysis.append({
                'Filename': filename,
                'Overall Confidence': round(overall_confidence, 2),
                'Fields Found': sum(1 for f in fields.values() if f['found']),
                'Total Fields': len(fields),
                'Success Rate': round((sum(1 for f in fields.values() if f['found']) / len(fields)) * 100, 1),
                'High Confidence Fields': sum(1 for f in fields.values() if f.get('confidence', 0) > 0.8),
                'Low Confidence Fields': sum(1 for f in fields.values() if f.get('confidence', 0) < 0.5 and f['found'])
            })
        
        return confidence_analysis
    
    def format_enhanced_excel_sheets(self, writer, main_df, summary_df, confidence_df):
        """Format Excel sheets with enhanced styling"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Colors for different confidence levels
            high_confidence_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
            medium_confidence_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")  # Amber
            low_confidence_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")  # Red
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Format main data sheet with confidence color coding
            ws_main = writer.sheets['Insurance Data']
            
            # Header formatting
            for cell in ws_main[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Color-code confidence values
            for row in range(2, ws_main.max_row + 1):
                for col in range(1, ws_main.max_column + 1):
                    cell = ws_main.cell(row=row, column=col)
                    
                    # Check if this is a confidence column
                    if 'Confidence)' in str(ws_main.cell(row=1, column=col).value):
                        try:
                            confidence_value = float(cell.value)
                            if confidence_value >= 0.8:
                                cell.fill = high_confidence_fill
                                cell.font = Font(color="FFFFFF", bold=True)
                            elif confidence_value >= 0.5:
                                cell.fill = medium_confidence_fill
                                cell.font = Font(color="000000", bold=True)
                            elif confidence_value > 0:
                                cell.fill = low_confidence_fill
                                cell.font = Font(color="FFFFFF", bold=True)
                        except (ValueError, TypeError):
                            pass
            
            # Auto-adjust column widths
            for column in ws_main.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws_main.column_dimensions[column_letter].width = adjusted_width
            
            # Format summary sheet
            ws_summary = writer.sheets['Summary']
            for cell in ws_summary[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Format confidence analysis sheet
            ws_confidence = writer.sheets['Confidence Analysis']
            for cell in ws_confidence[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust widths for all sheets
            for sheet in [ws_summary, ws_confidence]:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 30)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            self.logger.warning(f"Could not format Excel sheets: {e}")


# Maintain backward compatibility
class InsuranceExtractor(EnhancedInsuranceExtractor):
    """Backward compatibility wrapper"""
    pass


def main():
    """Test the enhanced insurance extractor"""
    extractor = EnhancedInsuranceExtractor()
    
    # Sample insurance text for testing with various formats
    sample_text = """
    MOTOR INSURANCE CERTIFICATE
    
    Policy Number: ABC123/2024/001
    Insured Person: Mr. John Doe
    Insurance Company: XYZ General Insurance Company Ltd
    Engine Number: ENG123456
    Chassis No.: CHS789012345
    
    PREMIUM BREAKDOWN:
    Own Damage Premium     Rs. 15,000.00
    Third Party Premium    Rs. 2,500.00
    Basic Premium Total    Rs. 17,500.00
    GST @ 18%             Rs. 3,150.00
    Total Amount Payable  Rs. 20,650.00
    
    Payment Information:
    Cheque Number: 123456
    Date of Payment: 15/01/2024
    Bank: State Bank of India
    
    Vehicle Information:
    Make & Model: Honda City ZX
    Body Type: Sedan
    """
    
    results = extractor.extract_insurance_data(sample_text, "test_insurance.pdf")
    
    print("Enhanced Extraction Results:")
    print("=" * 50)
    for field_key, field_info in results.items():
        status = "✅" if field_info['found'] else "❌"
        confidence = f"({field_info['confidence']:.2f})" if field_info['found'] else ""
        method = f"[{field_info['method']}]" if field_info['found'] else ""
        print(f"{status} {field_info['field_name']}: {field_info['value']} {confidence} {method}")

if __name__ == "__main__":
    main() 